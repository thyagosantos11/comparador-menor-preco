
import re
import threading
import os
import sys
import json
from pathlib import Path
from tkinter import filedialog, messagebox
import tkinter as tk

import customtkinter as ctk
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Tema ─────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

AZUL_ESCURO  = "#1F4E79"
AZUL_MED     = "#2E75B6"
VERDE        = "#1D9E75"
VERDE_CLARO  = "#C6EFCE"
CINZA_BG     = "#F4F7FB"
CINZA_CARD   = "#FFFFFF"
CINZA_BORDA  = "#D0DAE8"
TEXTO_ESCURO = "#1A2332"
TEXTO_MEDIO  = "#4A6080"
TEXTO_CLARO  = "#8099B3"

# ── Persistência de dados (produtos e fornecedores) ───────────────────────────
DATA_FILE = Path.home() / ".comparador_pedidos.json"

def carregar_dados_pedido():
    if DATA_FILE.exists():
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"produtos": [], "fornecedores": []}

def salvar_dados_pedido(dados):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

# ── Lógica de comparação ──────────────────────────────────────────────────────
INVALIDOS = {"f", "F", "?"}
COR_HEADER    = "1F4E79"
COR_MENOR     = "C6EFCE"
COR_LINHA_PAR = "EEF3FB"

def limpar_preco(valor):
    if pd.isna(valor): return float("nan")
    s = str(valor).strip()
    if s in INVALIDOS or s == "": return float("nan")
    s = re.sub(r"\s+", "", s)
    try: return float(s)
    except ValueError: return float("nan")

def montar_nome_produto(df_raw, fornecedores):
    col_produto = df_raw.columns[1]
    nomes, prefixo = [], ""
    for _, row in df_raw.iterrows():
        nome = str(row[col_produto]).strip() if pd.notna(row[col_produto]) else ""
        if not nome or nome == "nan":
            nomes.append(None); continue
        fornecedor_cols = df_raw.columns[3:3+len(fornecedores)]
        sem_preco = all(pd.isna(row[c]) or str(row[c]).strip() in INVALIDOS | {""} for c in fornecedor_cols)
        if nome.endswith(":") and sem_preco:
            prefixo = nome + " "; nomes.append(None)
        else:
            venda = row[df_raw.columns[0]]
            eh_subitem = pd.isna(venda) and len(nome.split()) <= 2 and prefixo
            nomes.append(prefixo + nome if eh_subitem else nome)
            if not eh_subitem: prefixo = ""
    return nomes

def carregar_planilha(arquivo):
    df_raw = pd.read_excel(arquivo, header=0)
    # Detecta automaticamente se a última coluna é "obs" ou um fornecedor
    ultima_col = str(df_raw.columns[-1]).strip().lower()
    nomes_obs = {"obs", "observacao", "observacoes", "observacao", "observacoes", "nan", "unnamed"}
    eh_obs = ultima_col in nomes_obs or ultima_col.startswith("unnamed")
    if not eh_obs:
        # Se nenhum valor numérico na última coluna, trata como obs
        valores_ultima = pd.to_numeric(df_raw.iloc[:, -1], errors="coerce")
        eh_obs = valores_ultima.notna().sum() == 0
    if eh_obs:
        fornecedores = list(df_raw.columns[3:-1])
        df_raw.columns = ["venda", "produto_orig", "quant"] + fornecedores + ["obs"]
    else:
        fornecedores = list(df_raw.columns[3:])
        df_raw.columns = ["venda", "produto_orig", "quant"] + fornecedores
    df_raw = df_raw.dropna(how="all")
    df_raw["produto"] = montar_nome_produto(df_raw, fornecedores)
    df_raw = df_raw[df_raw["produto"].notna()].reset_index(drop=True)
    for f in fornecedores:
        df_raw[f] = df_raw[f].apply(limpar_preco)
    df_raw["quant"] = df_raw["quant"].astype(str).str.strip().replace("nan", "—")
    df_raw["venda"] = pd.to_numeric(df_raw["venda"], errors="coerce")
    return df_raw[["produto", "quant", "venda"] + fornecedores], fornecedores

def calcular_menor(df, fornecedores):
    precos = df[fornecedores]
    sem_preco = precos.isna().all(axis=1)
    df["menor_preco"] = precos.min(axis=1)
    melhor = []
    for i, row in precos.iterrows():
        melhor.append("—" if sem_preco[i] else row.idxmin(skipna=True))
    df["melhor_forn"] = melhor
    df.loc[sem_preco, "menor_preco"] = float("nan")
    # Calcula % de lucro: (venda - menor_preco) / menor_preco * 100
    def calc_lucro(row):
        v = row["venda"]
        m = row["menor_preco"]
        if pd.isna(v) or pd.isna(m) or m == 0:
            return float("nan")
        return (v - m) / m * 100
    df["lucro_pct"] = df.apply(calc_lucro, axis=1)
    return df

def borda(cor="CCCCCC"):
    s = Side(style="thin", color=cor)
    return Border(left=s, right=s, top=s, bottom=s)

def cab(cell, texto, cor_bg=COR_HEADER):
    cell.value = texto
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=cor_bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borda("FFFFFF")

def cel(cell, valor, cor_fundo, bold=False, cor_fonte="000000", h_align="left", numero=False):
    cell.value = valor
    cell.font = Font(bold=bold, name="Arial", size=10, color=cor_fonte)
    cell.fill = PatternFill("solid", start_color=cor_fundo)
    cell.alignment = Alignment(horizontal=h_align, vertical="center")
    cell.border = borda()
    if numero and isinstance(valor, float) and not pd.isna(valor):
        cell.number_format = 'R$ #,##0.00'

def gerar_excel_comparacao(df, fornecedores, arquivo_saida):
    df.to_excel(arquivo_saida, index=False)
    wb = load_workbook(arquivo_saida)
    ws = wb.active
    ws.title = "Menor Preço por Produto"
    ws.delete_rows(1, ws.max_row)
    colunas = ["Produto", "Quant.", "Venda (R$)"] + [f.capitalize() for f in fornecedores] + ["Menor Preço", "Melhor Fornecedor", "% Lucro"]
    for ci, titulo in enumerate(colunas, 1):
        cab(ws.cell(row=1, column=ci), titulo)
    ws.row_dimensions[1].height = 30
    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        fundo = COR_LINHA_PAR if ri % 2 == 0 else "FFFFFF"
        menor = row["menor_preco"]
        melhor = row["melhor_forn"]
        venda = row["venda"]
        lucro = row["lucro_pct"]

        cel(ws.cell(ri, 1), row["produto"], fundo, bold=True)
        cel(ws.cell(ri, 2), row["quant"], fundo, h_align="center")

        # Coluna Venda
        val_venda = venda if pd.notna(venda) else "—"
        cel(ws.cell(ri, 3), val_venda, fundo, h_align="center", numero=True)

        for ci, forn in enumerate(fornecedores, start=4):
            preco = row[forn]
            eh_menor = pd.notna(preco) and pd.notna(menor) and abs(preco - menor) < 0.001
            fg = COR_MENOR if eh_menor else fundo
            cor_txt = "276221" if eh_menor else "000000"
            val = preco if pd.notna(preco) else "—"
            cel(ws.cell(ri, ci), val, fg, bold=eh_menor, cor_fonte=cor_txt, h_align="center", numero=True)

        ci_menor = len(fornecedores) + 4
        val_menor = menor if pd.notna(menor) else "—"
        cel(ws.cell(ri, ci_menor), val_menor, COR_MENOR, bold=True, cor_fonte="276221", h_align="center", numero=True)
        cel(ws.cell(ri, ci_menor + 1), melhor.capitalize() if melhor != "—" else "—", fundo, h_align="center")

        # Coluna % Lucro com cor condicional
        ci_lucro = ci_menor + 2
        if pd.notna(lucro):
            if lucro < 20:
                cor_lucro_bg = "FDECEA"
                cor_lucro_txt = "A32D2D"
            else:
                cor_lucro_bg = "C6EFCE"
                cor_lucro_txt = "276221"
            c = ws.cell(ri, ci_lucro)
            c.value = lucro / 100
            c.font = Font(bold=True, name="Arial", size=10, color=cor_lucro_txt)
            c.fill = PatternFill("solid", start_color=cor_lucro_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda()
            c.number_format = '0.00%'
        else:
            cel(ws.cell(ri, ci_lucro), "—", fundo, h_align="center")

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 13
    for ci in range(4, len(colunas) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "D2"
    wb.save(arquivo_saida)

def gerar_excel_pedido(produtos_selecionados, fornecedor_nome, pasta_saida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"
    colunas = ["Produto", "Unidade", "Quantidade", "Preço Unitário (R$)", "Observações"]
    for ci, titulo in enumerate(colunas, 1):
        cab(ws.cell(row=1, column=ci), titulo)
    ws.row_dimensions[1].height = 28
    for ri, prod in enumerate(produtos_selecionados, start=2):
        fundo = COR_LINHA_PAR if ri % 2 == 0 else "FFFFFF"
        cel(ws.cell(ri, 1), prod["nome"], fundo, bold=True)
        cel(ws.cell(ri, 2), prod["unidade"], fundo, h_align="center")
        cel(ws.cell(ri, 3), prod["quantidade"], fundo, h_align="center")
        cel(ws.cell(ri, 4), "", fundo, h_align="center")
        cel(ws.cell(ri, 5), "", fundo)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 25
    ws.freeze_panes = "A2"
    from datetime import date
    data_str = date.today().strftime("%d-%m-%Y")
    nome_arquivo = f"Pedido_{fornecedor_nome.replace(' ', '_')}_{data_str}.xlsx"
    caminho = Path(pasta_saida) / nome_arquivo
    wb.save(caminho)
    return caminho

# ── Interface Principal ───────────────────────────────────────────────────────

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Comparador de Menor Preço")
        self.geometry("720x760")
        self.resizable(False, False)
        self.configure(fg_color=CINZA_BG)
        self.arquivo_entrada = None
        self.arquivo_saida   = None
        self._build_ui()

    def _build_ui(self):
        # Header
        header = ctk.CTkFrame(self, fg_color=AZUL_ESCURO, corner_radius=0, height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(
            header, text="  📊  Comparador de Menor Preço",
            font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
            text_color="white",
        ).pack(side="left", padx=24)
        ctk.CTkLabel(
            header, text="Mercadinho Esperança",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color="#A8C8E8",
        ).pack(side="right", padx=24)

        # Abas
        self.tabview = ctk.CTkTabview(self, fg_color=CINZA_BG, segmented_button_fg_color=CINZA_CARD,
                                       segmented_button_selected_color=AZUL_ESCURO,
                                       segmented_button_selected_hover_color=AZUL_MED,
                                       segmented_button_unselected_color=CINZA_CARD,
                                       segmented_button_unselected_hover_color="#E0E8F5",
                                       text_color=TEXTO_ESCURO, text_color_disabled=TEXTO_CLARO)
        self.tabview.pack(fill="both", expand=True, padx=16, pady=(12, 16))
        self.tabview.add("📋  Montar Pedido")
        self.tabview.add("📊  Comparar Preços")

        self._build_aba_pedido(self.tabview.tab("📋  Montar Pedido"))
        self._build_aba_comparacao(self.tabview.tab("📊  Comparar Preços"))

    # ── ABA PEDIDO ────────────────────────────────────────────────────────────

    def _build_aba_pedido(self, parent):
        self._dados = carregar_dados_pedido()
        self._qtds = {}   # id_produto -> StringVar quantidade

        frame = ctk.CTkScrollableFrame(parent, fg_color=CINZA_BG)
        frame.pack(fill="both", expand=True)
        self._pedido_frame = frame

        self._render_pedido()

    def _render_pedido(self):
        for w in self._pedido_frame.winfo_children():
            w.destroy()

        frame = self._pedido_frame

        # ── Seção Produtos ──
        self._section_header(frame, "Produtos cadastrados")

        if not self._dados["produtos"]:
            ctk.CTkLabel(frame, text="Nenhum produto cadastrado ainda.",
                font=ctk.CTkFont(family="Segoe UI", size=12),
                text_color=TEXTO_CLARO).pack(anchor="w", padx=4, pady=4)
        else:
            grid = ctk.CTkFrame(frame, fg_color="transparent")
            grid.pack(fill="x", pady=(0, 8))
            for i, prod in enumerate(self._dados["produtos"]):
                pid = prod["id"]
                if pid not in self._qtds:
                    self._qtds[pid] = tk.StringVar(value="1")
                self._produto_row(grid, prod, i)

        # Formulário adicionar produto
        add_card = ctk.CTkFrame(frame, fg_color=CINZA_CARD, corner_radius=10, border_width=1, border_color=CINZA_BORDA)
        add_card.pack(fill="x", pady=(0, 16))
        inner = ctk.CTkFrame(add_card, fg_color="transparent")
        inner.pack(fill="x", padx=14, pady=12)
        ctk.CTkLabel(inner, text="Adicionar produto",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color=TEXTO_ESCURO).pack(anchor="w", pady=(0, 8))
        row = ctk.CTkFrame(inner, fg_color="transparent")
        row.pack(fill="x")
        self._entry_prod_nome = ctk.CTkEntry(row, placeholder_text="Nome do produto", width=280,
            font=ctk.CTkFont(family="Segoe UI", size=12))
        self._entry_prod_nome.pack(side="left")
        self._entry_prod_unidade = ctk.CTkEntry(row, placeholder_text="Unidade (kg, cx, un...)", width=150,
            font=ctk.CTkFont(family="Segoe UI", size=12))
        self._entry_prod_unidade.pack(side="left", padx=8)
        ctk.CTkButton(row, text="+ Adicionar", width=110, height=34,
            fg_color=AZUL_ESCURO, hover_color=AZUL_MED, corner_radius=8,
            font=ctk.CTkFont(family="Segoe UI", size=12),
            command=self._adicionar_produto).pack(side="left")

        # ── Seção Fornecedores ──
        self._section_header(frame, "Fornecedores")

        forn_card = ctk.CTkFrame(frame, fg_color=CINZA_CARD, corner_radius=10, border_width=1, border_color=CINZA_BORDA)
        forn_card.pack(fill="x", pady=(0, 16))
        forn_inner = ctk.CTkFrame(forn_card, fg_color="transparent")
        forn_inner.pack(fill="x", padx=14, pady=12)

        if self._dados["fornecedores"]:
            tags_frame = ctk.CTkFrame(forn_inner, fg_color="transparent")
            tags_frame.pack(fill="x", pady=(0, 10))
            for forn in self._dados["fornecedores"]:
                self._forn_tag(tags_frame, forn)

        frow = ctk.CTkFrame(forn_inner, fg_color="transparent")
        frow.pack(fill="x")
        self._entry_forn_nome = ctk.CTkEntry(frow, placeholder_text="Nome do fornecedor", width=300,
            font=ctk.CTkFont(family="Segoe UI", size=12))
        self._entry_forn_nome.pack(side="left")
        ctk.CTkButton(frow, text="+ Adicionar", width=110, height=34,
            fg_color=AZUL_ESCURO, hover_color=AZUL_MED, corner_radius=8,
            font=ctk.CTkFont(family="Segoe UI", size=12),
            command=self._adicionar_fornecedor).pack(side="left", padx=8)

        # ── Botão Gerar Pedido ──
        self._section_header(frame, "Gerar pedido")
        btn_frame = ctk.CTkFrame(frame, fg_color=CINZA_CARD, corner_radius=10, border_width=1, border_color=CINZA_BORDA)
        btn_frame.pack(fill="x", pady=(0, 16))
        btn_inner = ctk.CTkFrame(btn_frame, fg_color="transparent")
        btn_inner.pack(fill="x", padx=14, pady=14)
        ctk.CTkLabel(btn_inner,
            text="Selecione os produtos acima (marque a caixa), defina as quantidades e clique em Gerar.",
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color=TEXTO_CLARO, wraplength=580, justify="left").pack(anchor="w", pady=(0, 10))
        ctk.CTkButton(btn_inner, text="📥  Gerar planilhas de pedido",
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            height=44, fg_color=VERDE, hover_color="#0F6E56", corner_radius=10,
            command=self._gerar_pedido).pack(fill="x")

    def _section_header(self, parent, texto):
        ctk.CTkLabel(parent, text=texto.upper(),
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            text_color=TEXTO_CLARO).pack(anchor="w", padx=2, pady=(10, 4))

    def _produto_row(self, parent, prod, idx):
        fundo = CINZA_CARD
        card = ctk.CTkFrame(parent, fg_color=fundo, corner_radius=8, border_width=1, border_color=CINZA_BORDA)
        card.pack(fill="x", pady=3)
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=12, pady=8)

        # Checkbox seleção
        var = tk.BooleanVar(value=prod.get("selecionado", False))
        prod["_var"] = var
        ctk.CTkCheckBox(inner, text="", variable=var, width=24,
            checkbox_width=20, checkbox_height=20,
            fg_color=AZUL_ESCURO, hover_color=AZUL_MED,
            command=lambda p=prod, v=var: self._toggle_produto(p, v)
        ).pack(side="left")

        # Nome e unidade
        ctk.CTkLabel(inner, text=prod["nome"],
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            text_color=TEXTO_ESCURO).pack(side="left", padx=(8, 4))
        ctk.CTkLabel(inner, text=f"({prod['unidade']})",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=TEXTO_CLARO).pack(side="left")

        # Quantidade
        ctk.CTkLabel(inner, text="Qtd:",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=TEXTO_MEDIO).pack(side="left", padx=(16, 4))

        pid = prod["id"]
        qty_frame = ctk.CTkFrame(inner, fg_color="transparent")
        qty_frame.pack(side="left")

        def dec(p=prod):
            v = int(self._qtds[p["id"]].get() or 1)
            if v > 1:
                self._qtds[p["id"]].set(str(v - 1))

        def inc(p=prod):
            v = int(self._qtds[p["id"]].get() or 1)
            self._qtds[p["id"]].set(str(v + 1))

        ctk.CTkButton(qty_frame, text="−", width=28, height=28,
            fg_color=CINZA_BG, hover_color=CINZA_BORDA,
            text_color=TEXTO_ESCURO, corner_radius=6,
            font=ctk.CTkFont(size=14),
            command=dec).pack(side="left")

        ctk.CTkEntry(qty_frame, textvariable=self._qtds[pid], width=48, height=28,
            font=ctk.CTkFont(family="Segoe UI", size=12),
            justify="center").pack(side="left", padx=3)

        ctk.CTkButton(qty_frame, text="+", width=28, height=28,
            fg_color=CINZA_BG, hover_color=CINZA_BORDA,
            text_color=TEXTO_ESCURO, corner_radius=6,
            font=ctk.CTkFont(size=14),
            command=inc).pack(side="left")

        # Botão remover
        ctk.CTkButton(inner, text="✕", width=28, height=28,
            fg_color="transparent", hover_color="#FDECEA",
            text_color="#C0392B", corner_radius=6,
            font=ctk.CTkFont(size=13),
            command=lambda p=prod: self._remover_produto(p["id"])).pack(side="right")

    def _forn_tag(self, parent, forn):
        tag = ctk.CTkFrame(parent, fg_color="#EEF3FB", corner_radius=20)
        tag.pack(side="left", padx=(0, 6), pady=2)
        ctk.CTkLabel(tag, text=forn["nome"],
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=AZUL_ESCURO).pack(side="left", padx=(10, 4), pady=4)
        ctk.CTkButton(tag, text="✕", width=20, height=20,
            fg_color="transparent", hover_color="#D0DAE8",
            text_color=TEXTO_CLARO, corner_radius=10,
            font=ctk.CTkFont(size=11),
            command=lambda f=forn: self._remover_fornecedor(f["id"])).pack(side="left", padx=(0, 6))

    def _toggle_produto(self, prod, var):
        prod["selecionado"] = var.get()

    def _adicionar_produto(self):
        nome = self._entry_prod_nome.get().strip()
        unidade = self._entry_prod_unidade.get().strip() or "un"
        if not nome:
            messagebox.showwarning("Atenção", "Informe o nome do produto.")
            return
        novo = {"id": str(len(self._dados["produtos"])) + nome[:4], "nome": nome, "unidade": unidade, "selecionado": False}
        self._dados["produtos"].append(novo)
        self._qtds[novo["id"]] = tk.StringVar(value="1")
        salvar_dados_pedido(self._dados)
        self._render_pedido()

    def _remover_produto(self, pid):
        self._dados["produtos"] = [p for p in self._dados["produtos"] if p["id"] != pid]
        salvar_dados_pedido(self._dados)
        self._render_pedido()

    def _adicionar_fornecedor(self):
        nome = self._entry_forn_nome.get().strip()
        if not nome:
            messagebox.showwarning("Atenção", "Informe o nome do fornecedor.")
            return
        novo = {"id": str(len(self._dados["fornecedores"])) + nome[:4], "nome": nome}
        self._dados["fornecedores"].append(novo)
        salvar_dados_pedido(self._dados)
        self._render_pedido()

    def _remover_fornecedor(self, fid):
        self._dados["fornecedores"] = [f for f in self._dados["fornecedores"] if f["id"] != fid]
        salvar_dados_pedido(self._dados)
        self._render_pedido()

    def _gerar_pedido(self):
        selecionados = [p for p in self._dados["produtos"] if p.get("selecionado")]
        if not selecionados:
            messagebox.showwarning("Atenção", "Selecione ao menos um produto.")
            return
        if not self._dados["fornecedores"]:
            messagebox.showwarning("Atenção", "Cadastre ao menos um fornecedor.")
            return

        pasta = filedialog.askdirectory(title="Escolha a pasta para salvar os pedidos")
        if not pasta:
            return

        produtos_com_qtd = []
        for prod in selecionados:
            pid = prod["id"]
            try:
                qtd = int(self._qtds[pid].get())
            except Exception:
                qtd = 1
            produtos_com_qtd.append({
                "nome": prod["nome"],
                "unidade": prod["unidade"],
                "quantidade": qtd,
            })

        gerados = []
        for forn in self._dados["fornecedores"]:
            caminho = gerar_excel_pedido(produtos_com_qtd, forn["nome"], pasta)
            gerados.append(str(caminho))

        messagebox.showinfo(
            "Pedidos gerados!",
            f"{len(gerados)} planilha(s) gerada(s) em:\n{pasta}\n\n" + "\n".join(Path(g).name for g in gerados)
        )

        if sys.platform == "win32":
            os.startfile(pasta)
        else:
            os.system(f'xdg-open "{pasta}"')

    # ── ABA COMPARAÇÃO ────────────────────────────────────────────────────────

    def _build_aba_comparacao(self, parent):
        body = ctk.CTkFrame(parent, fg_color=CINZA_BG, corner_radius=0)
        body.pack(fill="both", expand=True, padx=8, pady=8)

        self._card_entrada(body)
        self._card_saida(body)

        self.btn_processar = ctk.CTkButton(
            body, text="⚡  Gerar Relatório Comparativo",
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            height=50, fg_color=AZUL_ESCURO, hover_color=AZUL_MED,
            corner_radius=10, command=self._iniciar_processamento, state="disabled",
        )
        self.btn_processar.pack(fill="x", pady=(4, 0))

        self.progress = ctk.CTkProgressBar(body, height=6, fg_color="#D0DAE8", progress_color=VERDE)
        self.progress.pack(fill="x", pady=(12, 0))
        self.progress.set(0)

        self.log_frame = ctk.CTkFrame(body, fg_color=CINZA_CARD, corner_radius=10, border_width=1, border_color=CINZA_BORDA)
        self.log_frame.pack(fill="both", expand=True, pady=(12, 0))
        ctk.CTkLabel(self.log_frame, text="Log de execução",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            text_color=TEXTO_CLARO).pack(anchor="w", padx=14, pady=(10, 0))
        self.log_box = ctk.CTkTextbox(self.log_frame,
            font=ctk.CTkFont(family="Consolas", size=12),
            fg_color=CINZA_CARD, text_color=TEXTO_MEDIO,
            border_width=0, activate_scrollbars=True, state="disabled")
        self.log_box.pack(fill="both", expand=True, padx=8, pady=(4, 10))

        self.btn_abrir = ctk.CTkButton(body, text="📂  Abrir arquivo gerado",
            font=ctk.CTkFont(family="Segoe UI", size=13),
            height=42, fg_color=VERDE, hover_color="#0F6E56",
            corner_radius=10, command=self._abrir_resultado, state="disabled")
        self.btn_abrir.pack(fill="x", pady=(10, 0))

    def _card_entrada(self, parent):
        card = ctk.CTkFrame(parent, fg_color=CINZA_CARD, corner_radius=10, border_width=1, border_color=CINZA_BORDA)
        card.pack(fill="x", pady=(0, 12))
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=16, pady=14)
        ctk.CTkLabel(inner, text="1. Selecione a planilha de entrada",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(inner, text="Arquivo RELAÇÃO DE COMPRA.xlsx",
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color=TEXTO_CLARO).pack(anchor="w", pady=(1, 10))
        row = ctk.CTkFrame(inner, fg_color="transparent")
        row.pack(fill="x")
        self.lbl_entrada = ctk.CTkLabel(row, text="Nenhum arquivo selecionado",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=TEXTO_CLARO, fg_color="#F0F4F9", corner_radius=6, anchor="w")
        self.lbl_entrada.pack(side="left", fill="x", expand=True, ipady=8, ipadx=10)
        ctk.CTkButton(row, text="Procurar...",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            width=100, height=36, fg_color=AZUL_ESCURO, hover_color=AZUL_MED,
            corner_radius=6, command=self._selecionar_entrada).pack(side="left", padx=(8, 0))

    def _card_saida(self, parent):
        card = ctk.CTkFrame(parent, fg_color=CINZA_CARD, corner_radius=10, border_width=1, border_color=CINZA_BORDA)
        card.pack(fill="x", pady=(0, 12))
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=16, pady=14)
        ctk.CTkLabel(inner, text="2. Escolha onde salvar o resultado",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color=TEXTO_ESCURO).pack(anchor="w")
        ctk.CTkLabel(inner, text="O arquivo resultado_menor_preco.xlsx será salvo aqui",
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color=TEXTO_CLARO).pack(anchor="w", pady=(1, 10))
        row = ctk.CTkFrame(inner, fg_color="transparent")
        row.pack(fill="x")
        self.lbl_saida = ctk.CTkLabel(row, text="Mesma pasta do arquivo de entrada",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=TEXTO_CLARO, fg_color="#F0F4F9", corner_radius=6, anchor="w")
        self.lbl_saida.pack(side="left", fill="x", expand=True, ipady=8, ipadx=10)
        ctk.CTkButton(row, text="Escolher...",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            width=100, height=36, fg_color="#5A7A9A", hover_color="#4A6A8A",
            corner_radius=6, command=self._selecionar_saida).pack(side="left", padx=(8, 0))

    def _selecionar_entrada(self):
        path = filedialog.askopenfilename(
            title="Selecionar planilha de entrada",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])
        if not path: return
        self.arquivo_entrada = path
        self.lbl_entrada.configure(text=f"✓  {Path(path).name}", text_color=VERDE)
        saida = str(Path(path).parent / "resultado_menor_preco.xlsx")
        self.arquivo_saida = saida
        self.lbl_saida.configure(text=f"  {Path(path).parent}", text_color=TEXTO_MEDIO)
        self.btn_processar.configure(state="normal")
        self.btn_abrir.configure(state="disabled")
        self.progress.set(0)
        self._log_clear()

    def _selecionar_saida(self):
        path = filedialog.asksaveasfilename(
            title="Salvar resultado como", defaultextension=".xlsx",
            initialfile="resultado_menor_preco.xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path: return
        self.arquivo_saida = path
        self.lbl_saida.configure(text=f"  {Path(path).parent}", text_color=TEXTO_MEDIO)

    def _iniciar_processamento(self):
        if not self.arquivo_entrada or not self.arquivo_saida: return
        self.btn_processar.configure(state="disabled", text="Processando...")
        self.btn_abrir.configure(state="disabled")
        self.progress.set(0)
        self._log_clear()
        threading.Thread(target=self._processar, daemon=True).start()

    def _processar(self):
        try:
            self._log("📂 Lendo planilha...")
            self.progress.set(0.15)
            df, fornecedores = carregar_planilha(self.arquivo_entrada)
            self._log(f"✓  {len(df)} produtos carregados")
            self._log(f"   Fornecedores detectados: {', '.join(f.capitalize() for f in fornecedores)}")
            self.progress.set(0.40)
            self._log("\n⚙️  Calculando menores preços...")
            df = calcular_menor(df, fornecedores)
            sem_cot = int(df["melhor_forn"].eq("—").sum())
            com_cot = len(df) - sem_cot
            self._log(f"✓  {com_cot} produtos com cotação  |  {sem_cot} sem cotação")
            lucro_baixo = int((df["lucro_pct"] < 20).sum()) if "lucro_pct" in df.columns else 0
            if lucro_baixo:
                self._log(f"⚠️  {lucro_baixo} produto(s) com margem abaixo de 20%")
            self.progress.set(0.65)
            self._log("\n💾 Gerando Excel formatado...")
            gerar_excel_comparacao(df, fornecedores, self.arquivo_saida)
            self.progress.set(0.95)
            self._log("\n🏆 Ranking de fornecedores:")
            contagem = df[df["melhor_forn"] != "—"]["melhor_forn"].value_counts()
            for forn, qtd in contagem.items():
                bar = "█" * min(int(qtd / max(contagem) * 20), 20)
                self._log(f"   {forn.capitalize():<14} {bar} {qtd} itens")
            self.progress.set(1.0)
            self._log(f"\n✅ Arquivo salvo em:\n   {self.arquivo_saida}")
            self.after(0, lambda: self.btn_abrir.configure(state="normal"))
        except Exception as e:
            self._log(f"\n❌ Erro: {e}")
            self.progress.set(0)
        finally:
            self.after(0, lambda: self.btn_processar.configure(
                state="normal", text="⚡  Gerar Relatório Comparativo"))

    def _abrir_resultado(self):
        if self.arquivo_saida and os.path.exists(self.arquivo_saida):
            os.startfile(self.arquivo_saida) if sys.platform == "win32" else os.system(f'xdg-open "{self.arquivo_saida}"')

    def _log(self, msg):
        def _do():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _do)

    def _log_clear(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")


if __name__ == "__main__":
    app = App()
    app.mainloop()
